#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Adapter for Clinical Data
临床数据的 Excel 适配器

Handles Excel format files (.xlsx, .xls).
"""

from pathlib import Path
from typing import List, Dict, Any, Optional

from .base import BaseAdapter

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


class ExcelAdapter(BaseAdapter):
    """
    Adapter for Excel format (.xlsx, .xls).
    Excel 格式的适配器。

    Features:
    - Multiple sheet support
    - Data type preservation
    - Formula evaluation (basic)
    - .xlsx and .xls support
    """

    def __init__(self, sheet_name: Optional[str] = None):
        """
        Initialize Excel adapter.

        Args:
            sheet_name: Name of sheet to load (None for first sheet)
        """
        super().__init__()
        self.sheet_name = sheet_name

    def load(
        self,
        path: str,
        header_row: int = 0
    ) -> List[Dict[str, Any]]:
        """
        Load data from Excel file.

        Args:
            path: Path to input file
            header_row: Row number containing headers (0-indexed)

        Returns:
            List of data records

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If Excel format is invalid or library not installed
        """
        p = self.validate_path(path, must_exist=True)

        if p.suffix.lower() == ".xlsx":
            return self._load_xlsx(path, header_row)
        elif p.suffix.lower() == ".xls":
            return self._load_xls(path, header_row)
        else:
            raise ValueError(f"Unsupported Excel format: {p.suffix}")

    def _load_xlsx(
        self,
        path: str,
        header_row: int
    ) -> List[Dict[str, Any]]:
        """Load .xlsx file using openpyxl."""
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for .xlsx files. "
                "Install it with: pip install openpyxl"
            )

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            # Select sheet
            if self.sheet_name:
                if self.sheet_name not in wb.sheetnames:
                    raise ValueError(f"Sheet not found: {self.sheet_name}")
                sheet = wb[self.sheet_name]
            else:
                sheet = wb.active

            # Read headers
            headers = None
            data = []

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx == header_row:
                    headers = [str(cell) if cell is not None else f"col_{i}"
                              for i, cell in enumerate(row)]
                elif row_idx > header_row:
                    if headers is None:
                        raise ValueError(f"No headers found at row {header_row}")

                    record = {}
                    for col_idx, value in enumerate(row):
                        key = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
                        record[key] = value

                    data.append(record)

            wb.close()
            return data

        except Exception as e:
            raise ValueError(f"Failed to read .xlsx file: {e}")

    def _load_xls(
        self,
        path: str,
        header_row: int
    ) -> List[Dict[str, Any]]:
        """Load .xls file using xlrd."""
        if not HAS_XLRD:
            raise ImportError(
                "xlrd is required for .xls files. "
                "Install it with: pip install xlrd"
            )

        try:
            workbook = xlrd.open_workbook(path)

            # Select sheet
            if self.sheet_name:
                if self.sheet_name not in workbook.sheet_names():
                    raise ValueError(f"Sheet not found: {self.sheet_name}")
                sheet = workbook.sheet_by_name(self.sheet_name)
            else:
                sheet = workbook.sheet_by_index(0)

            # Read data
            headers = None
            data = []

            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)

                if row_idx == header_row:
                    headers = [str(cell) if cell != "" else f"col_{i}"
                              for i, cell in enumerate(row)]
                elif row_idx > header_row:
                    if headers is None:
                        raise ValueError(f"No headers found at row {header_row}")

                    record = {}
                    for col_idx, value in enumerate(row):
                        key = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
                        record[key] = value

                    data.append(record)

            return data

        except Exception as e:
            raise ValueError(f"Failed to read .xls file: {e}")

    def save(
        self,
        data: List[Dict[str, Any]],
        path: str,
        sheet_name: str = "Sheet1"
    ) -> None:
        """
        Save data to Excel file (.xlsx).

        Args:
            data: List of data records to save
            path: Path to output file
            sheet_name: Name of sheet to create

        Raises:
            ValueError: If data format is invalid or library not installed
            IOError: If file cannot be written
        """
        self.validate_data(data)

        if not data:
            raise ValueError("Cannot save empty data to Excel")

        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for writing .xlsx files. "
                "Install it with: pip install openpyxl"
            )

        self.validate_path(path, must_exist=False)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Get headers from first record
            headers = list(data[0].keys())

            # Write header row
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # Write data rows
            for row_idx, record in enumerate(data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    value = record.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(path)

        except Exception as e:
            raise IOError(f"Failed to write Excel file: {e}")
